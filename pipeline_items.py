#!/usr/bin/env python3
"""
单据级提取管线

每页可能含多个单据(记账凭证/发票/银行回单), 每个单据输出一行Excel。
列: 文件路径 | 文件名 | 页码 | 单据序号 | 类型 | 凭证编号 | 记账日期 | 公司名称 |
     发票号码 | 开票日期 | 销售方名称 | 价税合计 | 回单编号 | 交易日期 | 交易金额 |
     备注(含摘要/科目等)

用法:
  python pipeline_items.py --dirs ./examples/pdfs --output output_items
"""
import io, json, os, re, sys, time
from pathlib import Path
from urllib.parse import quote

import fitz
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---- 配置 ----
# PyInstaller 路径适配
if getattr(sys, 'frozen', False):
    _BASE_DIR = Path(sys._MEIPASS)
else:
    _BASE_DIR = Path(__file__).parent


def _load_env():
    env = {}
    env_file = _BASE_DIR / ".env"
    if env_file.exists():
        for line in open(env_file, encoding="utf-8").readlines():
            line = line.strip()
            if line.startswith("#") or "=" not in line: continue
            k, v = line.split("=", 1); env[k.strip()] = v.strip()
    return env

ENV = _load_env()
APP_ID = os.environ.get("TEXTIN_APP_ID", ENV.get("TEXTIN_APP_ID", ""))
SECRET = os.environ.get("TEXTIN_SECRET_CODE", ENV.get("TEXTIN_SECRET_CODE", ""))
DS_KEY = os.environ.get("DEEPSEEK_API_KEY", ENV.get("DEEPSEEK_API_KEY", ""))

PDF_MD_URL = "https://api.textin.com/ai/service/v1/pdf_to_markdown"
DS_URL = "https://api.deepseek.com/chat/completions"


def validate_api_config():
    missing = []
    if not APP_ID:
        missing.append("TEXTIN_APP_ID")
    if not SECRET:
        missing.append("TEXTIN_SECRET_CODE")
    if not DS_KEY:
        missing.append("DEEPSEEK_API_KEY")
    if missing:
        names = ", ".join(missing)
        raise RuntimeError(
            f"缺少 API 配置: {names}. 请复制 .env.example 为 .env，或通过环境变量设置。"
        )

# ---- Prompt: 提取每页所有单据 ----
SYSTEM_PROMPT = """你是财务文档提取专家。请从单页markdown中提取该页包含的**所有**财务单据。

一页可能包含:
- 1个或多个记账凭证(有"借方金额"/"贷方金额"表格)
- 1张或多张发票(有"发票号码"/"开票日期")
- 1张或多张银行回单(有"回单编号"/"交易流水号")
- 也可能没有任何单据(审批单/附件等)

对每个找到的单据,返回其类型和所有可提取的字段。

返回JSON数组格式:
[
  {"类型":"记账凭证","凭证编号":"...","记账日期":"...","公司名称":"...","摘要":"...","金额":"..."},
  {"类型":"发票","发票号码":"...","开票日期":"...","销售方名称":"...","价税合计":"..."},
  {"类型":"银行回单","回单编号":"...","交易日期":"...","交易金额":"..."}
]

规则:
- 记账凭证: 凭证编号是markdown中独立的3-6位数字(位于日期和"1/1"附近); 记账日期在"日期"后; 公司名称在"编制单位"或"核算单位"后; 摘要取主要事项描述; 金额取表格最后一行的借方金额合计数字(去掉逗号和货币符号,只保留数字和小数点)。
- 发票: 发票号码在"发票号码:"后; 开票日期在"开票日期:"后; 销售方名称在销售方"名称:"后(NOT购买方); 价税合计在"价税合计"区域¥后。
- 银行回单: 回单编号在"回单编号:"后; 交易日期在"起息日:"或银行区域; 交易金额在"交易金额:"后。

如果没有任何单据,返回空数组[]。只返回JSON数组,不要解释。"""


def deepseek_extract_items(md_text, page_num, total_pages):
    validate_api_config()
    user_prompt = f"第{page_num}页(共{total_pages}页):\n\n{md_text[:7000]}"

    t0 = time.time()
    resp = requests.post(DS_URL, json={
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": 1000, "temperature": 0,
    }, headers={"Authorization": f"Bearer {DS_KEY}", "Content-Type": "application/json"}, timeout=90)
    elapsed = time.time() - t0
    data = resp.json()
    raw = data["choices"][0]["message"]["content"]
    tokens = data.get("usage", {})

    # 解析 JSON 数组
    items = []
    try:
        arr_match = re.search(r'\[.*\]', raw, re.DOTALL)
        if arr_match:
            items = json.loads(arr_match.group())
    except json.JSONDecodeError:
        # Try single object
        try:
            m = re.search(r'\{[^{}]*\}', raw, re.DOTALL)
            if m: items = [json.loads(m.group())]
        except: pass

    return items, elapsed, tokens, raw


def process_pdf(pdf_path, output_base, verbose=True):
    """返回: list of item dicts, 每个 dict 对应一个单据行"""
    validate_api_config()
    pdf_path = Path(pdf_path)
    pdf_name = pdf_path.stem
    # 保存绝对路径，确保之后能找到 PDF
    rel_path = str(pdf_path.resolve())

    pdf_dir = output_base / pdf_name
    pdf_dir.mkdir(parents=True, exist_ok=True)
    page_dir = pdf_dir / "pages"
    page_dir.mkdir(exist_ok=True)

    doc = fitz.open(pdf_path)
    total_pages = len(doc)

    all_items = []  # 所有单据行
    all_raw = []    # 保存原始数据

    if verbose:
        print(f"  {pdf_name} ", end="", flush=True)

    for page_idx in range(total_pages):
        page_num = page_idx + 1

        # 单页 PDF
        single = fitz.open(); single.insert_pdf(doc, from_page=page_idx, to_page=page_idx)
        buf = io.BytesIO(); single.save(buf); single.close()

        # pdf_to_markdown
        t0 = time.time()
        resp = requests.post(f"{PDF_MD_URL}?parse_mode=vlm&table_flavor=html",
                             data=buf.getvalue(),
                             headers={"x-ti-app-id": APP_ID, "x-ti-secret-code": SECRET,
                                      "Content-Type": "application/octet-stream"},
                             timeout=60)
        md_time = time.time() - t0
        data = resp.json()
        if data.get("code") != 200:
            if verbose: print(f"p{page_num}:ERR ", end="", flush=True)
            continue
        md_text = data.get("result", {}).get("markdown", "")
        if not md_text: continue

        # 保存 markdown
        (page_dir / f"p{page_num}_markdown.md").write_text(md_text, encoding="utf-8")

        # DeepSeek 提取
        items, llm_time, tokens, raw = deepseek_extract_items(md_text, page_num, total_pages)

        # 保存 DeepSeek 响应
        (page_dir / f"p{page_num}_deepseek.json").write_text(
            json.dumps({"page": page_num, "raw": raw, "items": items, "tokens": tokens},
                       ensure_ascii=False, indent=2), encoding="utf-8")

        # 每个单据一行
        for seq, item in enumerate(items):
            if not isinstance(item, dict): continue
            doc_type = item.get("类型", "")

            row = {
                "文件路径": rel_path,
                "文件名": pdf_name,
                "页码": page_num,
                "单据序号": seq + 1,
                "单据类型": doc_type,
                "凭证编号": "",
                "记账日期": "",
                "公司名称": "",
                "摘要": "",
                "金额": "",
                "发票号码": "",
                "开票日期": "",
                "销售方名称": "",
                "价税合计": "",
                "回单编号": "",
                "交易日期": "",
                "交易金额": "",
                "备注": "",
            }

            if doc_type == "记账凭证":
                row["凭证编号"] = str(item.get("凭证编号", "")).strip()
                row["记账日期"] = str(item.get("记账日期", "")).strip()
                row["公司名称"] = str(item.get("公司名称", "")).strip()
                row["摘要"] = str(item.get("摘要", "")).strip()[:200]
                row["金额"] = str(item.get("金额", "")).strip()

            elif doc_type == "发票":
                row["发票号码"] = str(item.get("发票号码", "")).strip()
                row["开票日期"] = str(item.get("开票日期", "")).strip()
                row["销售方名称"] = str(item.get("销售方名称", "")).strip()
                row["价税合计"] = str(item.get("价税合计", "")).strip()

            elif doc_type == "银行回单":
                row["回单编号"] = str(item.get("回单编号", "")).strip()
                row["交易日期"] = str(item.get("交易日期", "")).strip()
                row["交易金额"] = str(item.get("交易金额", "")).strip()

            else:
                # 未知类型, 把原始item放入备注
                row["备注"] = json.dumps(item, ensure_ascii=False)
                row["单据类型"] = doc_type or "未知"

            all_items.append(row)

        if verbose:
            n_items = len(items)
            types = ",".join(it.get("类型","?")[:2] for it in items) if items else "空"
            print(f"p{page_num}:{n_items}单({types}) ", end="", flush=True)

    doc.close()

    # 保存完整中间数据
    (pdf_dir / "intermediate.json").write_text(
        json.dumps({"pdf_name": pdf_name, "total_pages": total_pages,
                    "total_items": len(all_items), "items": all_items, "raw_pages": all_raw},
                   ensure_ascii=False, indent=2), encoding="utf-8")

    if verbose:
        print(f"→ {len(all_items)}个单据", flush=True)

    return all_items


# ---- Excel 输出 ----
def write_items_excel(all_rows, output_dir, base_dir=None):
    """所有单据写入一个 Excel"""
    if base_dir is None:
        base_dir = Path.cwd()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "单据明细"

    HEADERS = [
        "文件路径", "文件名", "页码", "单据序号", "单据类型",
        "凭证编号", "记账日期", "公司名称", "摘要", "金额",
        "发票号码", "开票日期", "销售方名称", "价税合计",
        "回单编号", "交易日期", "交易金额",
        "备注",
    ]

    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    # Type-specific fills
    type_fills = {
        "记账凭证": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
        "发票": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "银行回单": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    }

    # Header
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h); c.font = hfont; c.fill = hf; c.border = thin

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(all_rows)+1}"

    # Data rows
    for ri, row in enumerate(all_rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            v = row.get(h, "")

            # 文件名列(B) 加 PDF 超链接 (HYPERLINK公式)
            if h == "文件名" and row.get("文件路径", ""):
                pdf_abs = str((base_dir / row["文件路径"]).resolve())
                # 转义路径中的特殊字符，用HYPERLINK公式
                display = v.replace('"', '""')
                c = ws.cell(row=ri, column=ci, value=f'=HYPERLINK("{pdf_abs}","{display}")')
                c.font = Font(color="0563C1", underline="single")
            else:
                c = ws.cell(row=ri, column=ci, value=v)

            c.border = thin
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Row color by type
        doc_type = row.get("单据类型", "")
        fill = type_fills.get(doc_type)
        if fill:
            for ci in range(1, len(HEADERS)+1):
                ws.cell(row=ri, column=ci).fill = fill

    # Column widths
    widths = {
        1: 50, 2: 30, 3: 6, 4: 6, 5: 10,
        6: 14, 7: 14, 8: 30, 9: 40, 10: 14,
        11: 25, 12: 14, 13: 35, 14: 16,
        15: 25, 16: 14, 17: 16,
        18: 50,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    fp = output_dir / "单据明细.xlsx"
    wb.save(fp)
    return fp


# ---- Main ----
def main():
    import argparse
    parser = argparse.ArgumentParser(description="单据级智能提取")
    parser.add_argument("--dirs", nargs="*", required=True, help="一个或多个包含 PDF 的目录")
    parser.add_argument("--output", default="output_items")
    parser.add_argument("--verbose", type=int, default=1)
    args = parser.parse_args()

    base = _BASE_DIR
    output_dir = base / args.output; output_dir.mkdir(parents=True, exist_ok=True)

    # 收集 PDF
    pdf_files = []
    for d in args.dirs:
        dp = Path(d).expanduser()
        if not dp.is_absolute():
            dp = base / dp
        if dp.exists():
            pdf_files.extend(sorted(dp.rglob("*.pdf")))
    if not pdf_files:
        print("无PDF!"); return

    print(f"{'='*70}")
    print(f"  单据级智能提取 — 每页所有单据独立成行")
    print(f"  PDF: {len(pdf_files)} 份")
    print(f"{'='*70}")

    all_rows = []
    t_start = time.time()
    ok, err = 0, 0

    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"[{i}/{len(pdf_files)}] ", end="", flush=True)
        try:
            rows = process_pdf(pdf_path, output_dir, verbose=args.verbose >= 1)
            all_rows.extend(rows)
            ok += 1
        except Exception as e:
            print(f"❌ {e}")
            err += 1

    # 输出 Excel
    fp = write_items_excel(all_rows, output_dir, base)
    total_t = time.time() - t_start

    # 统计
    types = {}
    for r in all_rows:
        t = r.get("单据类型", "未知")
        types[t] = types.get(t, 0) + 1

    print(f"\n{'='*70}")
    print(f"  完成! {ok}成功 {err}失败 | {total_t/60:.1f}分钟")
    print(f"  单据总数: {len(all_rows)}")
    for t, n in sorted(types.items()):
        print(f"    {t}: {n}")
    print(f"\n  输出: {fp}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
