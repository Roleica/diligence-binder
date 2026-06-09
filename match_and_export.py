#!/usr/bin/env python3
"""Match reviewed vouchers, invoices, and bank receipts into an Excel workbook."""

import csv
import re
import sys
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CATEGORY_SUFFIX_RE = re.compile(r"^(\d+)(销特|销|研特|研|管特|管)$")
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
EXPENSE_CATEGORIES = (
    ("销", "销售费用"),
    ("研", "研发费用"),
    ("管", "管理费用"),
)


def clean_amount(val):
    if not val:
        return None
    s = str(val).strip()
    s = re.sub(r"^[$￥A-Z]{1,4}\s*", "", s).replace(",", "").replace(" ", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def amounts_match(a, b, tolerance=0.5):
    if a is None or b is None or a == 0 or b == 0:
        return False
    diff = abs(a - b)
    return diff < tolerance or diff / max(a, b) < 0.01


def parse_expected_voucher_no(pdf_name):
    """Parse voucher number from names like 24060123销 -> 123."""
    match = CATEGORY_SUFFIX_RE.match(pdf_name)
    if not match:
        return ""

    digits = match.group(1)
    if len(digits) >= 7:
        return digits[4:].lstrip("0")

    for vlen in (4, 3):
        if len(digits) >= vlen + 4:
            return digits[-vlen:].lstrip("0")
    return ""


def has_amount_match(voucher, invoices, banks):
    amount = voucher.get("_金额")
    if amount is None:
        return False
    return any(amounts_match(amount, inv.get("_价税合计")) for inv in invoices) or any(
        amounts_match(amount, bank.get("_交易金额")) for bank in banks
    )


def is_foreign_inv(inv):
    seller = str(inv.get("销售方名称", "") or "")
    inv_no = str(inv.get("发票号码", "") or "")
    inv_date = str(inv.get("开票日期", "") or "")
    if re.search(r"[A-Za-z]{5,}", seller):
        return True
    if re.search(r"[A-Za-z]", inv_no) and not CJK_RE.search(inv_no):
        return True
    return bool(re.search(r"[A-Za-z]", inv_date))


def is_foreign_bank(bank):
    amount = str(bank.get("交易金额", "") or "").upper()
    return any(marker in amount for marker in ("$", "USD", "AUD", "EUR"))


def split_items_by_type(items):
    return (
        [i for i in items if i["type"] == "记账凭证"],
        [i for i in items if i["type"] == "发票"],
        [i for i in items if i["type"] == "银行回单"],
    )


def drop_blank_documents(invoices, banks):
    invoices = [i for i in invoices if i.get("发票号码", "").strip() or i.get("价税合计", "").strip()]
    banks = [b for b in banks if b.get("回单编号", "").strip() or b.get("交易金额", "").strip()]
    return invoices, banks


def dedupe_invoices(invoices):
    seen = set()
    deduped = []
    for invoice in invoices:
        key = invoice.get("发票号码", "")
        if key and key in seen:
            continue
        seen.add(key)
        deduped.append(invoice)
    return deduped


def build_fake_reasons(fake_vouchers, boundary, expected_vno):
    reasons = []
    for voucher in fake_vouchers:
        vno = voucher.get("凭证编号", "")
        if voucher["page"] >= boundary and boundary < 999:
            reasons.append(vno + "(页码异常)")
        elif expected_vno and vno and vno != expected_vno:
            reasons.append(vno + "(编号≠" + expected_vno + ")")
        else:
            reasons.append(vno)
    return reasons


def expense_category(pdf_name):
    for marker, category in EXPENSE_CATEGORIES:
        if marker in pdf_name:
            return category
    return ""


def read_csv(csv_path):
    groups = defaultdict(list)
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pdf_name = row.get('文件名', '').strip()
            if not pdf_name: continue
            doc_type = row.get('单据类型', '').strip()
            item = {
                'pdf_name': pdf_name,
                'page': int(row.get('页码', 0) or 0),
                'type': doc_type,
                '凭证编号': row.get('凭证编号', '').strip(),
                '记账日期': row.get('记账日期', '').strip(),
                '公司名称': row.get('公司名称', '').strip(),
                '摘要': row.get('摘要', '').strip(),
                '金额': row.get('金额', '').strip(),
                '发票号码': row.get('发票号码', '').strip(),
                '开票日期': row.get('开票日期', '').strip(),
                '销售方名称': row.get('销售方名称', '').strip(),
                '价税合计': row.get('价税合计', '').strip(),
                '回单编号': row.get('回单编号', '').strip(),
                '交易日期': row.get('交易日期', '').strip(),
                '交易金额': row.get('交易金额', '').strip(),
                '_金额': clean_amount(row.get('金额', '')),
                '_价税合计': clean_amount(row.get('价税合计', '')),
                '_交易金额': clean_amount(row.get('交易金额', '')),
            }
            groups[pdf_name].append(item)
    return groups


def match_items(items):
    """新规则匹配"""
    all_vouchers, invoices, banks = split_items_by_type(items)

    # === 规则1: 剔除假凭证 (页码顺序 + 文件名凭证号) ===
    first_inv_page = min((i['page'] for i in invoices), default=999)
    first_bank_page = min((i['page'] for i in banks), default=999)
    boundary = min(first_inv_page, first_bank_page)

    pdf_name = items[0]['pdf_name'] if items else ''
    expected_vno = parse_expected_voucher_no(pdf_name)

    real_v, fake_v = [], []
    for v in all_vouchers:
        vno = str(v.get('凭证编号', '')).strip().lstrip('0')
        # 页码规则: 在第一张发票/回单之前的页面 = 真
        is_real_by_page = v['page'] < boundary
        # 文件名规则: 凭证号匹配文件名 = 真
        is_real_by_name = expected_vno and vno and vno == expected_vno
        # 平级: 任一规则认为真就保留; 两个都认为是假的才剔除
        if is_real_by_page or is_real_by_name:
            real_v.append(v)
        else:
            fake_v.append(v)

    # 去重: 匹配文件名的凭证号出现多个 → 保留金额和回单/发票匹配的, 否则保留页码最前的
    best_by_vno = {}
    for v in real_v:
        vno = str(v.get('凭证编号', '')).strip().lstrip('0')
        if not (vno and expected_vno and vno == expected_vno):
            continue
        if vno not in best_by_vno:
            best_by_vno[vno] = v
            continue
        # 同名凭证号: 比较哪个更好
        old_v = best_by_vno[vno]
        old_match = has_amount_match(old_v, invoices, banks)
        new_match = has_amount_match(v, invoices, banks)
        if new_match and not old_match:
            fake_v.append(old_v); best_by_vno[vno] = v
        elif old_match and not new_match:
            fake_v.append(v)
        elif v['page'] < old_v['page']:
            fake_v.append(old_v); best_by_vno[vno] = v
        else:
            fake_v.append(v)
    real_v = [v for v in real_v if v not in fake_v]

    if not real_v and all_vouchers:
        real_v = [all_vouchers[0]]

    vouchers = real_v
    # 清理空条目(LLM幻觉产生的空白行)
    invoices, banks = drop_blank_documents(invoices, banks)
    invoices = dedupe_invoices(invoices)

    matched, used_v, used_i, used_b = [], set(), set(), set()

    for vi, v in enumerate(vouchers):
        if vi in used_v: continue
        v_amt = v['_金额']
        best_i, best_b = None, None

        # 找匹配发票
        for ii, inv in enumerate(invoices):
            if ii in used_i: continue
            if v_amt and amounts_match(v_amt, inv['_价税合计']):
                best_i = ii; break
        # 找匹配回单
        for bi, bank in enumerate(banks):
            if bi in used_b: continue
            if v_amt and amounts_match(v_amt, bank['_交易金额']):
                best_b = bi; break
        # 交叉匹配
        if best_i is None and best_b is not None:
            b_amt = banks[best_b]['_交易金额']
            for ii, inv in enumerate(invoices):
                if ii in used_i: continue
                if amounts_match(b_amt, inv['_价税合计']):
                    best_i = ii; break
        if best_b is None and best_i is not None:
            i_amt = invoices[best_i]['_价税合计']
            for bi, bank in enumerate(banks):
                if bi in used_b: continue
                if amounts_match(i_amt, bank['_交易金额']):
                    best_b = bi; break

        # === 规则2: 多发票合并 ===
        if best_b is not None and best_i is None:
            target = banks[best_b]['_交易金额'] or v_amt
            if target:
                avail = [(ii, inv) for ii, inv in enumerate(invoices) if ii not in used_i and inv['_价税合计']]
                for r in range(2, min(len(avail)+1, 6)):
                    for combo in combinations(avail, r):
                        if amounts_match(sum(inv['_价税合计'] for _, inv in combo), target):
                            first_ii, first_inv = combo[0]
                            best_i = first_ii
                            nums = [inv['发票号码'] for _, inv in combo]
                            invoices[first_ii]['发票号码'] = nums[0] + '等'
                            invoices[first_ii]['_价税合计'] = target
                            invoices[first_ii]['价税合计'] = str(target)
                            for oi, _ in combo[1:]:
                                used_i.add(oi)
                            break
                    if best_i is not None: break

        # === 规则3: 外币/外国发票 ===
        # 凭证和回单匹配, 发票不匹配 → 找外币发票
        if best_b is not None and best_i is None:
            for ii, inv in enumerate(invoices):
                if ii in used_i: continue
                if is_foreign_inv(inv):
                    best_i = ii
                    break

        # 凭证和发票匹配, 回单不匹配 → 找外币回单
        if best_i is not None and best_b is None:
            for bi, bank in enumerate(banks):
                if bi in used_b: continue
                if is_foreign_bank(bank):
                    best_b = bi
                    break

        # 全外币: 1凭证+1票+1单都存在, 但金额都不匹配, 全标记外币直接配对
        if best_i is None and best_b is None and len(vouchers) == 1 and len(invoices) == 1 and len(banks) == 1:
            if is_foreign_inv(invoices[0]) or is_foreign_bank(banks[0]):
                best_i = 0; best_b = 0

        used_v.add(vi)
        if best_i is not None: used_i.add(best_i)
        if best_b is not None: used_b.add(best_b)
        fake_reasons = build_fake_reasons(fake_v, boundary, expected_vno)

        matched.append({
            'voucher': v, 'fake_vouchers': fake_v,
            'fake_reasons': fake_reasons,
            'invoice': invoices[best_i] if best_i is not None else None,
            'bank': banks[best_b] if best_b is not None else None,
        })

    # 后处理: 未匹配的外币发票+回单配对 (可能凭证已被用作其他)
    for ii, inv in enumerate(invoices):
        if ii in used_i: continue
        if not is_foreign_inv(inv): continue
        for bi, bank in enumerate(banks):
            if bi in used_b: continue
            if not (is_foreign_bank(bank) or is_foreign_inv(inv)): continue
            # 找一个未用凭证, 或已用但无票无单的凭证
            best_v = None
            for vi, v in enumerate(vouchers):
                if vi not in used_v:
                    best_v = vi; break
            if best_v is None:
                for mi, m in enumerate(matched):
                    if m['voucher'] and not m['invoice'] and not m['bank']:
                        best_v = 'm'+str(mi); break
            if best_v is not None:
                if isinstance(best_v, int):
                    used_v.add(best_v)
                    v = vouchers[best_v]
                else:
                    mi = int(best_v[1:])
                    v = matched[mi]['voucher']
                    matched[mi]['voucher'] = None  # Remove from old row
                used_i.add(ii); used_b.add(bi)
                matched.append({'voucher': v, 'fake_vouchers': fake_v,
                    'invoice': inv, 'bank': bank, 'fake_reasons': []})
            break

    # 剩余
    for ii, inv in enumerate(invoices):
        if ii not in used_i:
            matched.append({'voucher': None, 'fake_vouchers': fake_v, 'invoice': inv, 'bank': None})
    for bi, bank in enumerate(banks):
        if bi not in used_b:
            found = False
            for m in matched:
                if m['bank'] is None and m['invoice'] is not None:
                    if amounts_match(bank['_交易金额'], m['invoice']['_价税合计']):
                        m['bank'] = bank; found = True; break
            if not found:
                matched.append({'voucher': None, 'fake_vouchers': fake_v, 'invoice': None, 'bank': bank})

    return matched


def write_gt_excel(all_matched, output_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "匹配结果"
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    sf = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sfont = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.merge_cells("A3:A3"); ws.merge_cells("D3:G3"); ws.merge_cells("H3:K3"); ws.merge_cells("L3:N3")
    for c, t in [(1,"文件路径"),(2,"主体"),(3,"样本索引"),(4,"记账凭证"),(8,"发票"),(12,"付款水单"),(15,"匹配"),(16,"费用"),(17,"备注")]:
        cl = ws.cell(row=3, column=c, value=t); cl.font = hfont; cl.fill = hf; cl.border = thin
    for c, t in {4:"编号",5:"日期",6:"摘要",7:"金额",8:"编号",9:"日期",10:"供应商",11:"金额",12:"编号",13:"日期",14:"金额"}.items():
        cl = ws.cell(row=4, column=c, value=t); cl.font = sfont; cl.fill = sf; cl.border = thin

    for ri, m in enumerate(all_matched):
        row = 5 + ri
        v = m['voucher'] or {}; i = m['invoice'] or {}; b = m['bank'] or {}
        company = v.get('公司名称','')
        pdf_name = v.get('pdf_name','') or i.get('pdf_name','') or b.get('pdf_name','')
        vals = {
            1: '', 2: company, 3: pdf_name,
            4: v.get('凭证编号',''), 5: v.get('记账日期',''), 6: v.get('摘要',''), 7: v.get('金额',''),
            8: i.get('发票号码',''), 9: i.get('开票日期',''), 10: i.get('销售方名称',''), 11: i.get('价税合计',''),
            12: b.get('回单编号',''), 13: b.get('交易日期',''), 14: b.get('交易金额',''), 15: '',
        }
        # 匹配状态
        v_amt, i_amt, b_amt = v.get('_金额'), i.get('_价税合计'), b.get('_交易金额')
        mismatch = False; is_merge = False; is_foreign = False
        amts = [a for a in [v_amt, i_amt, b_amt] if a is not None]
        for x in range(len(amts)):
            for y in range(x+1, len(amts)):
                if not amounts_match(amts[x], amts[y]):
                    mismatch = True
        if i.get('发票号码','') and '等' in str(i.get('发票号码','')): is_merge = True
        if any(c in str(i.get('价税合计',''))+str(b.get('交易金额','')) for c in ['$','USD','AUD','EUR']): is_foreign = True

        if is_merge: vals[15] = '🔗合并'
        elif is_foreign: vals[15] = '💱外币'
        elif mismatch: vals[15] = '⚠️金额不对'
        else: vals[15] = '✅'

        # 假凭证标记
        fake_info = '剔除:' + ', '.join(m.get('fake_reasons', [])) if m.get('fake_reasons') else ''
        vals[17] = fake_info

        for c, val in vals.items():
            ws.cell(row=row, column=c, value=val if val else '').border = thin
        if mismatch and not is_merge and not is_foreign:
            for c in range(1, 18): ws.cell(row=row, column=c).fill = yellow
        elif is_merge or is_foreign:
            for c in [7, 11, 14]: ws.cell(row=row, column=c).fill = blue
        if fake_info:
            ws.cell(row=row, column=2).fill = gray

        # 费用类型
        ws.cell(row=row, column=16, value=expense_category(pdf_name)).border = thin

    for c, w in {1:55, 2:30, 3:30, 4:12, 5:14, 6:45, 7:16, 8:28, 9:14, 10:38, 11:18, 12:28, 13:14, 14:18, 15:12, 16:12, 17:30}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) <= 1:
        print("用法: python match_and_export.py /path/to/审核结果.csv")
        sys.exit(1)
    csv_path = sys.argv[1]
    csv_path = Path(csv_path)
    if not csv_path.exists():
        print(f"文件不存在: {csv_path}"); sys.exit(1)

    print(f"读取: {csv_path}")
    groups = read_csv(csv_path)
    print(f"  {len(groups)} 份 PDF")

    all_matched = []
    for pdf_name in sorted(groups.keys()):
        items = groups[pdf_name]
        matched = match_items(items)
        all_matched.extend(matched)
        vc = sum(1 for i in items if i['type']=='记账凭证')
        ic = sum(1 for i in items if i['type']=='发票')
        bc = sum(1 for i in items if i['type']=='银行回单')
        fake = len(matched[0].get('fake_vouchers',[])) if matched else 0
        info = f" (剔{fake}假)" if fake else ""
        print(f"  {pdf_name}: {vc}证+{ic}票+{bc}单 → {len(matched)}条{info}")

    output = csv_path.parent / "匹配结果.xlsx"
    write_gt_excel(all_matched, output)
    total = len(all_matched)
    perfect = sum(1 for m in all_matched if m['voucher'] and m['invoice'] and m['bank'])
    print(f"\n✅ {output} — {total} 条 ({perfect} 完整)")


if __name__ == "__main__":
    main()
